# backend/app/api/v1/knowledge.py
import logging
import time
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, Body
from sqlalchemy.ext.asyncio import AsyncSession
import asyncio

from app.core.database import get_db, AsyncSessionLocal
from app.core.security import get_current_admin_id
from app.crud.knowledge import crud_knowledge_base, crud_knowledge_item
from app.schemas.knowledge import (
    KnowledgeBaseCreate,
    KnowledgeBaseUpdate,
    KnowledgeBaseOut,
    KnowledgeItemCreate,
    KnowledgeItemUpdate,
    KnowledgeItemOut,
    KnowledgeItemDetail,
    KnowledgeItemQuery,
    ImportProgress,
    SearchRequest,
    SearchResponse,
    SearchResultItem,
    SimilarQuestionOut,
)
from app.services.knowledge.vectorizer import (
    upsert_item_vectors,
    delete_item_vectors,
    search_similar,
)
from app.services.knowledge.importer import (
    parse_excel,
    run_import_task,
    get_progress,
    get_error_details,
    create_import_task_id,
)
from app.utils.response import Response, PageResponse

logger = logging.getLogger(__name__)
router = APIRouter()


# ==================== 辅助函数 ====================


def _tags_to_list(tags_jsonb) -> Optional[List[str]]:
    """JSONB tags转列表"""
    if not tags_jsonb:
        return None
    if isinstance(tags_jsonb, dict):
        return tags_jsonb.get("tags", [])
    if isinstance(tags_jsonb, list):
        return tags_jsonb
    return []


def _item_to_out(
    item, similar_count: int = 0, similar_questions: list = None
) -> KnowledgeItemOut:
    """KnowledgeItem模型 → KnowledgeItemOut Schema"""
    return KnowledgeItemOut(
        id=item.id,
        knowledge_base_id=item.knowledge_base_id,
        title=item.title,
        answer=item.answer_content or "",
        answer_type=item.answer_type,
        category=item.category,
        tags=_tags_to_list(item.tags),
        status=item.status,
        similar_count=similar_count,
        similar_questions=similar_questions or [],
        created_at=item.created_at,
        updated_at=item.updated_at,
    )


def _base_to_out(base, item_count: int = 0) -> KnowledgeBaseOut:
    """KnowledgeBase模型 → KnowledgeBaseOut Schema"""
    return KnowledgeBaseOut(
        id=base.id,
        name=base.name,
        description=base.description,
        status=getattr(base, "status", 1),
        item_count=item_count,
        created_at=base.created_at,
        updated_at=base.updated_at,
    )


# ==================== 知识库管理 ====================


@router.get("/bases", summary="获取知识库列表")
async def list_knowledge_bases(
    db: AsyncSession = Depends(get_db),
    admin_id: int = Depends(get_current_admin_id),
):
    bases = await crud_knowledge_base.get_list(db)
    result = []
    for base in bases:
        count = await crud_knowledge_base.get_item_count(db, base.id)
        result.append(_base_to_out(base, item_count=count))
    return Response.success(data=result)


@router.post("/bases", summary="创建知识库")
async def create_knowledge_base(
    obj_in: KnowledgeBaseCreate,
    db: AsyncSession = Depends(get_db),
    admin_id: int = Depends(get_current_admin_id),
):
    base = await crud_knowledge_base.create(db, obj_in)
    return Response.success(data=_base_to_out(base, item_count=0))


@router.put("/bases/{kb_id}", summary="更新知识库")
async def update_knowledge_base(
    kb_id: int,
    obj_in: KnowledgeBaseUpdate,
    db: AsyncSession = Depends(get_db),
    admin_id: int = Depends(get_current_admin_id),
):
    base = await crud_knowledge_base.get(db, kb_id)
    if not base:
        raise HTTPException(status_code=404, detail="知识库不存在")

    base = await crud_knowledge_base.update(db, kb_id, obj_in)
    count = await crud_knowledge_base.get_item_count(db, base.id)
    return Response.success(data=_base_to_out(base, item_count=count))


# ==================== 知识条目管理 ====================


@router.get("/bases/{kb_id}/items", summary="获取知识条目列表")
async def list_knowledge_items(
    kb_id: int,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    keyword: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    status: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    admin_id: int = Depends(get_current_admin_id),
):
    try:
        base = await crud_knowledge_base.get(db, kb_id)
        if not base:
            raise HTTPException(status_code=404, detail="知识库不存在")

        query = KnowledgeItemQuery(
            page=page,
            page_size=page_size,
            keyword=keyword,
            category=category,
            status=status,
        )
        items, total = await crud_knowledge_item.get_list(db, kb_id, query)

        result = []
        for item in items:
            try:
                count = await crud_knowledge_item.get_similar_count(db, item.id)
                sim_questions = (
                    [sq.question for sq in item.similar_questions]
                    if item.similar_questions
                    else []
                )
                result.append(
                    _item_to_out(
                        item, similar_count=count, similar_questions=sim_questions
                    )
                )
            except Exception as e:
                logger.error(f"处理知识条目 {item.id} 失败: {e}")
                result.append(_item_to_out(item, similar_count=0, similar_questions=[]))

        return PageResponse.success(
            data=result,
            total=total,
            page=page,
            page_size=page_size,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取知识条目列表失败: {e}")
        raise HTTPException(status_code=500, detail=f"获取列表失败: {str(e)}")


@router.post("/bases/{kb_id}/items", summary="创建单条知识条目")
async def create_knowledge_item(
    kb_id: int,
    obj_in: KnowledgeItemCreate,
    db: AsyncSession = Depends(get_db),
    admin_id: int = Depends(get_current_admin_id),
):
    base = await crud_knowledge_base.get(db, kb_id)
    if not base:
        raise HTTPException(status_code=404, detail="知识库不存在")

    item = await crud_knowledge_item.create(db, kb_id, obj_in)

    try:
        await upsert_item_vectors(
            item_id=item.id,
            kb_id=kb_id,
            title=obj_in.title,
            similar_questions=obj_in.similar_questions,
        )
    except Exception as e:
        logger.error(f"向量化失败 item_id={item.id}: {e}")

    similar_count = await crud_knowledge_item.get_similar_count(db, item.id)
    return Response.success(data=_item_to_out(item, similar_count=similar_count))


@router.get("/items/{item_id}", summary="获取知识条目详情")
async def get_knowledge_item(
    item_id: int,
    db: AsyncSession = Depends(get_db),
    admin_id: int = Depends(get_current_admin_id),
):
    item = await crud_knowledge_item.get_with_similars(db, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="知识条目不存在")

    similar_questions = [
        SimilarQuestionOut(
            id=sq.id,
            question=sq.question,
            created_at=sq.created_at,
        )
        for sq in item.similar_questions
    ]

    result = KnowledgeItemDetail(
        id=item.id,
        knowledge_base_id=item.knowledge_base_id,
        title=item.title,
        answer=item.answer_content or "",
        answer_type=item.answer_type,
        category=item.category,
        tags=_tags_to_list(item.tags),
        status=item.status,
        similar_count=len(similar_questions),
        created_at=item.created_at,
        updated_at=item.updated_at,
        similar_questions=similar_questions,
    )
    return Response.success(data=result)


@router.put("/items/{item_id}", summary="更新知识条目")
async def update_knowledge_item(
    item_id: int,
    obj_in: KnowledgeItemUpdate,
    db: AsyncSession = Depends(get_db),
    admin_id: int = Depends(get_current_admin_id),
):
    item = await crud_knowledge_item.get(db, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="知识条目不存在")

    updated = await crud_knowledge_item.update(db, item_id, obj_in)

    if obj_in.title or obj_in.similar_questions is not None:
        try:
            await delete_item_vectors(item_id)
            latest = await crud_knowledge_item.get_with_similars(db, item_id)
            if latest:
                await upsert_item_vectors(
                    item_id=item_id,
                    kb_id=latest.knowledge_base_id,
                    title=latest.title,
                    similar_questions=[sq.question for sq in latest.similar_questions],
                )
        except Exception as e:
            logger.error(f"重新向量化失败 item_id={item_id}: {e}")

    similar_count = await crud_knowledge_item.get_similar_count(db, item_id)
    return Response.success(data=_item_to_out(updated, similar_count=similar_count))


@router.delete("/items/{item_id}", summary="删除知识条目")
async def delete_knowledge_item(
    item_id: int,
    db: AsyncSession = Depends(get_db),
    admin_id: int = Depends(get_current_admin_id),
):
    item = await crud_knowledge_item.get(db, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="知识条目不存在")

    try:
        await delete_item_vectors(item_id)
    except Exception as e:
        logger.error(f"删除向量失败 item_id={item_id}: {e}")

    await crud_knowledge_item.delete(db, item_id)
    return Response.success(message="删除成功")


@router.post("/items/batch-delete", summary="批量删除知识条目")
async def batch_delete_items(
    item_ids: List[int] = Body(..., description="要删除的知识条目ID列表"),
    db: AsyncSession = Depends(get_db),
    admin_id: int = Depends(get_current_admin_id),
):
    if not item_ids:
        raise HTTPException(status_code=400, detail="ID列表不能为空")

    deleted_count = await crud_knowledge_item.delete_batch(db, item_ids)
    return Response.success(message=f"成功删除 {deleted_count} 条知识条目")


# ==================== Excel模板下载 ====================


@router.get("/bases/template", summary="下载Excel导入模板")
async def download_template(
    admin_id: int = Depends(get_current_admin_id),
):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "知识库导入模板"

    headers = [
        "知识ID(留空新增)",
        "分类",
        "知识标题",
        "相似问法(换行分隔)",
        "答案类型(0文本1HTML)",
        "答案内容",
        "标签(逗号分隔)",
    ]

    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws.append(headers)

    for col_idx, col in enumerate(ws.columns, start=1):
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = [18, 12, 25, 30, 18, 35, 20][
            col_idx - 1
        ]
        for cell in col:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

    example_data = [
        [
            "",
            "退款",
            "如何申请退款？",
            "怎么退款？\n退款流程是什么？\n我想退款怎么办？",
            "0",
            "请联系客服申请退款，审核通过后3个工作日内原路返回。",
            "退款,售后",
        ],
        [
            "",
            "物流",
            "物流什么时候到？",
            "怎么还没收到货？\n快递到哪了？",
            "0",
            "请提供订单号，我帮您查询物流信息。",
            "物流,快递",
        ],
    ]

    for row_data in example_data:
        ws.append(row_data)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.column in [4, 6]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.row_dimensions[1].height = 25
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 40

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=knowledge_import_template.xlsx"
        },
    )


# ==================== Excel导入 ====================


@router.post("/bases/{kb_id}/import", summary="Excel批量导入")
async def import_knowledge_excel(
    kb_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    admin_id: int = Depends(get_current_admin_id),
):
    base = await crud_knowledge_base.get(db, kb_id)
    if not base:
        raise HTTPException(status_code=404, detail="知识库不存在")

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持Excel文件")

    file_bytes = await file.read()
    if len(file_bytes) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="文件大小不能超过10MB")

    # 快速统计行数（用于返回给前端）
    import openpyxl
    from io import BytesIO

    wb_preview = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws_preview = wb_preview.active
    total_rows = 0
    for row in ws_preview.iter_rows(min_row=2, values_only=True):
        if not all(cell is None or str(cell).strip() == "" for cell in row):
            total_rows += 1

    task_id = create_import_task_id()

    asyncio.create_task(
        run_import_task(
            task_id=task_id,
            kb_id=kb_id,
            file_bytes=file_bytes,
            db_factory=AsyncSessionLocal,
        )
    )

    return Response.success(
        data={"task_id": task_id, "total": total_rows, "message": "导入任务已启动"}
    )


@router.get("/import/progress/{task_id}", summary="查询导入进度")
async def get_import_progress(
    task_id: str,
    admin_id: int = Depends(get_current_admin_id),
):
    progress = await get_progress(task_id)
    if not progress:
        raise HTTPException(status_code=404, detail="任务不存在或已过期")

    errors = await get_error_details(task_id)
    progress.has_errors = errors is not None and len(errors) > 0

    return Response.success(data=progress)


@router.get("/import/errors/{task_id}", summary="查询导入错误详情")
async def get_import_errors(
    task_id: str,
    admin_id: int = Depends(get_current_admin_id),
):
    errors = await get_error_details(task_id)
    if errors is None:
        raise HTTPException(status_code=404, detail="错误详情不存在或已过期")
    return Response.success(data={"errors": errors, "count": len(errors)})


@router.get("/import/download-result/{task_id}", summary="下载导入结果Excel")
async def download_import_result(
    task_id: str,
    admin_id: int = Depends(get_current_admin_id),
):
    from fastapi.responses import StreamingResponse
    from app.services.knowledge.importer import get_result_file
    from io import BytesIO

    data = await get_result_file(task_id)

    if not data:
        raise HTTPException(status_code=404, detail="结果文件不存在或已过期")

    return StreamingResponse(
        BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=import_result_{task_id[:8]}.xlsx",
        },
    )


# ==================== 检索效果测试 ====================


# ==================== 检索效果测试 ====================


@router.post("/bases/{kb_id}/search", summary="检索效果测试")
async def search_knowledge(
    kb_id: int,
    req: SearchRequest,
    db: AsyncSession = Depends(get_db),
    admin_id: int = Depends(get_current_admin_id),
):
    base = await crud_knowledge_base.get(db, kb_id)
    if not base:
        raise HTTPException(status_code=404, detail="知识库不存在")

    start_time = time.time()

    try:
        vector_hits = await search_similar(
            query=req.query,
            kb_id=kb_id,
            top_k=req.top_k,
            score_threshold=req.score_threshold,
        )
    except Exception as e:
        logger.error(f"向量检索失败: {e}")
        vector_hits = []

    item_score_map: dict = {}
    for hit in vector_hits:
        item_id = hit["item_id"]
        if (
            item_id not in item_score_map
            or hit["vector_score"] > item_score_map[item_id]["vector_score"]
        ):
            item_score_map[item_id] = hit

    results = []
    for item_id, hit in item_score_map.items():
        item = await crud_knowledge_item.get(db, item_id)
        if not item:
            continue
        results.append(
            SearchResultItem(
                item_id=item_id,
                title=item.title,
                answer=item.answer_content or "",
                answer_type=item.answer_type,
                category=item.category,
                score=hit["vector_score"],
                bm25_score=0.0,
                vector_score=hit["vector_score"],
                matched_question=hit["question"],
            )
        )

    results.sort(key=lambda x: x.score, reverse=True)
    results = results[: req.top_k]
    elapsed_ms = round((time.time() - start_time) * 1000, 2)

    return Response.success(
        data=SearchResponse(
            query=req.query,
            results=results,
            total=len(results),
            elapsed_ms=elapsed_ms,
        )
    )

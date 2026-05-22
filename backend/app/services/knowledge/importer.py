# backend/app/services/knowledge/importer.py
import asyncio
import logging
import uuid
from typing import List, Tuple, Dict, Any, Optional
from copy import copy
import openpyxl
from openpyxl.styles import PatternFill, Font
from io import BytesIO

from sqlalchemy.ext.asyncio import AsyncSession

from app.schemas.knowledge import (
    ImportRowData,
    ImportProgress,
    ImportResult,
    AnswerType,
)
from app.crud.knowledge import crud_knowledge_item
from app.services.knowledge.vectorizer import batch_upsert_vectors
import app.core.redis

logger = logging.getLogger(__name__)

# ==================== Redis进度存储Key ====================


def _progress_key(task_id: str) -> str:
    return f"import:progress:{task_id}"


def _errors_key(task_id: str) -> str:
    return f"import:errors:{task_id}"


def _result_key(task_id: str) -> str:
    return f"import:result:{task_id}"


# ==================== 进度管理 ====================


async def set_progress(
    task_id: str,
    status: str,
    total: int = 0,
    processed: int = 0,
    succeeded: int = 0,
    failed: int = 0,
    error_msg: Optional[str] = None,
    error_details: Optional[List[dict]] = None,
    has_result_file: bool = False,
):
    """将导入进度写入Redis"""
    progress = round(processed / total * 100, 1) if total > 0 else 0.0
    data = {
        "task_id": task_id,
        "status": status,
        "total": total,
        "processed": processed,
        "succeeded": succeeded,
        "failed": failed,
        "progress": progress,
        "error_msg": error_msg or "",
        "has_result_file": "1" if has_result_file else "0",
    }
    key = _progress_key(task_id)
    await app.core.redis.redis_client.hset(
        key, mapping={k: str(v) for k, v in data.items()}
    )
    await app.core.redis.redis_client.expire(key, 3600)

    if error_details:
        errors_key = _errors_key(task_id)
        import json

        await app.core.redis.redis_client.setex(
            errors_key, 3600, json.dumps(error_details[:100])
        )


async def set_result_file(task_id: str, excel_bytes: bytes):
    """存储带错误标注的结果Excel到Redis（二进制模式）"""
    redis = await app.core.redis.get_redis_binary()
    key = _result_key(task_id)
    await redis.setex(key, 3600, excel_bytes)


async def get_result_file(task_id: str) -> Optional[bytes]:
    """读取结果Excel bytes"""
    redis = await app.core.redis.get_redis_binary()
    key = _result_key(task_id)
    data = await redis.get(key)
    return data


async def get_progress(task_id: str) -> Optional[ImportProgress]:
    """从Redis读取导入进度"""
    key = _progress_key(task_id)
    data = await app.core.redis.redis_client.hgetall(key)
    if not data:
        return None

    return ImportProgress(
        task_id=data.get("task_id", task_id),
        status=data.get("status", "unknown"),
        total=int(data.get("total", 0)),
        processed=int(data.get("processed", 0)),
        succeeded=int(data.get("succeeded", 0)),
        failed=int(data.get("failed", 0)),
        progress=float(data.get("progress", 0.0)),
        error_msg=data.get("error_msg") or None,
    )


async def get_error_details(task_id: str) -> Optional[List[dict]]:
    """从Redis读取导入错误详情"""
    import json

    errors_key = _errors_key(task_id)
    data = await app.core.redis.redis_client.get(errors_key)
    if not data:
        return None
    try:
        return json.loads(data)
    except:
        return None


# ==================== Excel解析 ====================


def parse_excel(file_bytes: bytes) -> Tuple[List[ImportRowData], List[dict]]:
    """
    解析Excel文件，返回 (有效行列表, 失败行列表)

    Excel列定义：
    | A      | B    | C        | D          | E    | F    | G    |
    | 知识ID | 分类 | 知识标题 | 相似问法   | 答案类型 | 答案内容 | 标签 |

    答案类型：0=纯文本(默认), 1=富文本HTML
    相似问法：多个用换行符分隔
    """
    valid_rows: List[ImportRowData] = []
    failed_rows: List[dict] = []

    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        ws = wb.active

        logger.info(f"Excel解析: 总行数={ws.max_row}, 总列数={ws.max_column}")
        for i, row in enumerate(
            ws.iter_rows(min_row=1, max_row=3, values_only=True), start=1
        ):
            logger.info(f"Excel第{i}行内容: {row}")

    except Exception as e:
        logger.error(f"Excel文件解析失败: {e}")
        raise ValueError(f"Excel文件格式错误: {str(e)}")
    except Exception as e:
        logger.error(f"Excel文件解析失败: {e}")
        raise ValueError(f"Excel文件格式错误: {str(e)}")

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        try:
            title = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            answer = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ""
            answer_type_raw = (
                str(row[4]).strip() if len(row) > 4 and row[4] is not None else "0"
            )
            similar_raw = (
                str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
            )
            category = (
                str(row[1]).strip() if len(row) > 1 and row[1] is not None else None
            )
            tags = str(row[6]).strip() if len(row) > 6 and row[6] is not None else None

            # 校验必填字段
            if not title:
                failed_rows.append(
                    {
                        "row": row_idx,
                        "reason": "知识标题不能为空",
                        "data": str(row[:2]),
                    }
                )
                continue

            if not answer:
                failed_rows.append(
                    {
                        "row": row_idx,
                        "reason": "答案内容不能为空",
                        "data": title,
                    }
                )
                continue

            # 解析答案类型
            try:
                answer_type = AnswerType(int(answer_type_raw))
            except (ValueError, TypeError):
                answer_type = AnswerType.TEXT

            # 解析相似问法（换行分隔）
            similar_questions = []
            if similar_raw:
                for q in (
                    similar_raw.replace("\r\n", "\n").replace("\r", "\n").split("\n")
                ):
                    q = q.strip()
                    if q:
                        similar_questions.append(q)

            # 清理空值
            category = category if category else None
            tags = tags if tags else None

            valid_rows.append(
                ImportRowData(
                    title=title,
                    answer=answer,
                    answer_type=answer_type,
                    category=category,
                    tags=tags,
                    similar_questions=similar_questions,
                )
            )

        except Exception as e:
            logger.warning(f"第{row_idx}行解析失败: {e}")
            failed_rows.append(
                {
                    "row": row_idx,
                    "reason": str(e),
                    "data": str(row),
                }
            )

    logger.info(f"Excel解析完成，有效行: {len(valid_rows)}，失败行: {len(failed_rows)}")
    return valid_rows, failed_rows


# ==================== 导入任务执行 ====================


async def run_import_task(
    task_id: str,
    kb_id: int,
    file_bytes: bytes,
    db_factory,
    batch_size: int = 100,
):
    """
    异步执行导入任务（在后台Task中运行）

    流程：
    1. 解析Excel → 获取有效行列表
    2. 分批写入PostgreSQL
    3. 批量向量化写入Qdrant
    4. 错误标注回写Excel并存储到Redis
    """
    logger.info(f"导入任务开始 task_id={task_id}, kb_id={kb_id}")

    await set_progress(task_id, status="running", total=0)

    failed_rows: List[dict] = []
    failed_count = 0
    succeeded_count = 0

    try:
        # 加载原始Excel用于后续错误标注
        wb = openpyxl.load_workbook(BytesIO(file_bytes))
        ws = wb.active

        # 在最后添加"导入结果"列
        result_col = ws.max_column + 1
        header_cell = ws.cell(row=1, column=result_col)
        header_cell.value = "导入结果"
        header_cell.font = Font(bold=True, color="FFFFFF")
        header_cell.fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_cell.alignment = openpyxl.styles.Alignment(
            horizontal="center", vertical="center"
        )

        green_fill = PatternFill(
            start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"
        )
        red_fill = PatternFill(
            start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"
        )

        # ── Step1：解析Excel ──
        valid_rows, parse_failed = parse_excel(file_bytes)
        failed_rows.extend(parse_failed)

        for pf in parse_failed:
            row_num = pf.get("row", 0)
            if row_num >= 2:
                cell = ws.cell(row=row_num, column=result_col)
                cell.value = f"❌ 解析失败: {pf.get('reason', '未知错误')}"
                cell.fill = red_fill

        total = len(valid_rows)

        await set_progress(
            task_id,
            status="running",
            total=total,
            processed=0,
            succeeded=0,
            failed=len(failed_rows),
            error_details=parse_failed if parse_failed else None,
        )

        if total == 0:
            await set_progress(
                task_id,
                status="done",
                total=0,
                error_msg="Excel中没有有效数据行，请检查格式是否正确",
                error_details=parse_failed if parse_failed else None,
                has_result_file=True,
            )
            output = BytesIO()
            wb.save(output)
            await set_result_file(task_id, output.getvalue())
            return

        # ── Step2：分批写入PostgreSQL ──
        vector_items: List[Dict[str, Any]] = []
        row_idx = 2

        for batch_start in range(0, total, batch_size):
            batch = valid_rows[batch_start : batch_start + batch_size]

            try:
                async with db_factory() as db:
                    created_items = await crud_knowledge_item.create_batch(
                        db=db,
                        kb_id=kb_id,
                        items=batch,
                    )

                for item_schema, db_item in zip(batch, created_items):
                    vector_items.append(
                        {
                            "item_id": db_item.id,
                            "kb_id": kb_id,
                            "title": item_schema.title,
                            "similar_questions": item_schema.similar_questions,
                        }
                    )

                    cell = ws.cell(row=row_idx, column=result_col)
                    cell.value = "✅ 导入成功"
                    cell.fill = green_fill

                    succeeded_count += 1
                    row_idx += 1

            except Exception as e:
                logger.error(f"批次写入失败 batch_start={batch_start}: {e}")
                error_str = str(e)
                for row_item in batch:
                    cell = ws.cell(row=row_idx, column=result_col)
                    cell.value = f"❌ 数据库写入失败: {error_str[:200]}"
                    cell.fill = red_fill
                    failed_rows.append(
                        {
                            "row": row_idx,
                            "reason": f"数据库写入失败: {error_str}",
                            "data": row_item.title,
                        }
                    )
                    failed_count += 1
                    row_idx += 1

            processed = min(batch_start + batch_size, total)
            await set_progress(
                task_id,
                status="running",
                total=total,
                processed=processed,
                succeeded=succeeded_count,
                failed=len(failed_rows),
            )

        # ── Step3：批量向量化写入Qdrant ──
        if vector_items:
            logger.info(f"开始向量化，共{len(vector_items)}条")
            try:
                total_vectors = await batch_upsert_vectors(
                    items=vector_items,
                    batch_size=50,
                )
                logger.info(f"向量化完成，共存入{total_vectors}个向量")
            except Exception as e:
                logger.error(f"向量化失败: {e}")

        # ── Step4：保存带标注的结果Excel ──
        output = BytesIO()
        wb.save(output)
        result_bytes = output.getvalue()

        await set_result_file(task_id, result_bytes)

        # 标记完成
        await set_progress(
            task_id,
            status="done",
            total=total,
            processed=total,
            succeeded=succeeded_count,
            failed=len(failed_rows),
            error_details=failed_rows if failed_rows else None,
            has_result_file=len(failed_rows) > 0,
        )

        logger.info(
            f"导入任务完成 task_id={task_id}，成功={succeeded_count}，失败={len(failed_rows)}"
        )

    except Exception as e:
        logger.error(f"导入任务异常 task_id={task_id}: {e}")
        await set_progress(
            task_id,
            status="failed",
            error_msg=str(e),
        )


# ==================== 创建导入任务 ====================


def create_import_task_id() -> str:
    """生成唯一任务ID"""
    return f"import_{uuid.uuid4().hex}"
